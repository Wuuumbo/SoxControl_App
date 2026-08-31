#!/usr/bin/env python3
"""Initialise le dossier data/ a partir du classeur SOX_Controls_Workbook_2 FINAL.xlsm.

A executer une seule fois (ou pour reinitialiser les donnees de depart).
Necessite openpyxl : python -m pip install openpyxl
"""
import json
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
XLSM = ROOT.parent.parent / "SOX_Controls_Workbook_2 FINAL.xlsm"

CONTROLS = [
    {"ref": "C1C2", "nom": "Affiliate Current Accounts", "description": "Verification des comptes courants affilies", "contact": "Quantum"},
    {"ref": "C3", "nom": "Overdrafts Set-up", "description": "Mise en place des overdrafts", "contact": "Quantum"},
    {"ref": "N3", "nom": "Overdrafts Follow-up", "description": "Suivi quotidien des overdrafts", "contact": "-"},
    {"ref": "C4", "nom": "Affiliate Financing Set-up", "description": "Mise en place financements affilies", "contact": "Quantum"},
    {"ref": "C5", "nom": "Compliance Check", "description": "Verification de conformite", "contact": "Quantum", "note": "macro OK"},
    {"ref": "C6", "nom": "Contracts Bank Account Tracker", "description": "Suivi contrats comptes bancaires", "contact": "-"},
    {"ref": "C7", "nom": "Counterparty", "description": "Due diligence contreparties", "contact": "Quantum"},
    {"ref": "C8", "nom": "Audit Trails US", "description": "Pistes d'audit US", "contact": "-"},
    {"ref": "C9", "nom": "MarkitWire vs Quantum", "description": "Reconciliation MarkitWire / Quantum", "contact": "-"},
]

N3_COLUMNS = [
    ("id", "ID"), ("date_rapport", "Date Rapport"), ("run_for_date", "Run For Date"),
    ("affiliate", "Affiliate"), ("devise", "Devise"), ("limite_od", "Limite OD (M)"),
    ("utilisation", "Utilisation (M)"), ("depassement", "Depassement"),
    ("commentaire_eod", "Commentaire EoD"), ("valide_powerbi", "Valide Power BI"),
]

C5_COLUMNS = [
    ("deal_no", "Deal No."), ("ticket_number", "Ticket Number"), ("themis_ref", "THEMIS Ref"),
    ("internal_total_affiliate", "Internal Total Affiliate"), ("counterparty", "Counterparty"),
    ("country_of_payment", "Country of Payment"), ("beneficiary_bank_bic", "Beneficiary Bank BIC"),
    ("beneficiary_bank_name", "Beneficiary Bank Name"), ("intermediary_bic", "Intermediary BIC"),
    ("value_date", "Value Date"), ("ccy", "CCY"), ("amount", "Amount"), ("comments", "Comments"),
]

DEFAULT_PROCEDURE_STEPS = [
    {"etape": "1", "action": "Extraction des donnees", "mode_operatoire": "", "statut": "-", "date": "", "contact": ""},
    {"etape": "2", "action": "Tirage des echantillons", "mode_operatoire": "", "statut": "-", "date": "", "contact": ""},
    {"etape": "3", "action": "Analyse des echantillons", "mode_operatoire": "", "statut": "-", "date": "", "contact": ""},
    {"etape": "4", "action": "Conclusion et signature", "mode_operatoire": "", "statut": "-", "date": "", "contact": ""},
]

N3_PROCEDURE = [
    {"etape": "1", "action": "Extraction des donnees", "mode_operatoire": "Lancer l'extraction depuis le systeme source (rapport Counterparty Overdraft Limits + Power BI)", "statut": "-", "date": "", "contact": "Equipe BO"},
    {"etape": "2", "action": "Chargement SOX+", "mode_operatoire": "Charger le fichier ZIP dans SOX+ : l'outil genere automatiquement 2 dates aleatoires", "statut": "-", "date": "", "contact": "SOX+"},
    {"etape": "3", "action": "Envoi mail demande", "mode_operatoire": "Mail auto SOX+ ou rediger manuellement la demande de justificatifs pour les 2 dates", "statut": "-", "date": "", "contact": "Metier"},
    {"etape": "4", "action": "Analyse overdraft quotidien", "mode_operatoire": "Verifier le montant overdraft quotidien pour chaque date selectionnee", "statut": "-", "date": "", "contact": "Arthur (BO)"},
    {"etape": "5", "action": "Verification Power BI", "mode_operatoire": "Valider rapport Power BI : commentaires presents, depassements justifies, signatures", "statut": "-", "date": "", "contact": "Arthur (BO)"},
    {"etape": "6", "action": "Verification EoD", "mode_operatoire": "S'assurer que tout est integre dans le process de fin de journee (End of Day)", "statut": "-", "date": "", "contact": "Equipe BO"},
    {"etape": "7", "action": "Reprise modele", "mode_operatoire": "Reprendre le modele de controle du mois precedent : verifier dates, montants > 1M EUR commentes", "statut": "-", "date": "", "contact": "Controleur"},
    {"etape": "8", "action": "Reception reponse", "mode_operatoire": "Attendre la reponse metier avec justificatifs complets", "statut": "-", "date": "", "contact": "Arthur (BO)"},
    {"etape": "9", "action": "Signature et depot", "mode_operatoire": "Signer le document final et deposer dans SharePoint controle", "statut": "-", "date": "", "contact": "Controleur"},
]

PERIOD = "2026-07"

# Identite synthetique pour les donnees reprises de l'ancien classeur Excel
# (par opposition a une action realisee en direct dans l'outil, qui capture
# l'identite Windows reelle de la personne).
SEED_IDENTITY = {"display_name": "Import initial (ancien classeur Excel)", "username": None, "computer": None}


def write_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def cell_to_value(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def row_to_dict(row_values, columns):
    return {key: cell_to_value(v) for (key, _label), v in zip(columns, row_values)}


DEFAULT_PROCEDURE_COLUMNS = [
    ("etape", "Etape"), ("action", "Action"), ("mode_operatoire", "Mode operatoire"),
    ("statut", "Statut"), ("date", "Date"), ("contact", "Contact"),
]


def build_meta(c):
    return {
        "ref": c["ref"],
        "nom": c["nom"],
        "description": c["description"],
        "contact": c.get("contact", "-"),
        "note": c.get("note", ""),
        "contacts": [],
        "mail_templates": [],
    }


def seed_control(c, columns, rows, procedure_steps, echantillons_ids=None, date_test=""):
    ref = c["ref"]
    cdir = DATA / "controls" / ref
    write_json(cdir / "meta.json", build_meta(c))

    pdir = cdir / "periods" / PERIOD
    write_json(pdir / "raw_data.json", {
        "columns": [{"key": k, "label": l} for k, l in columns],
        "rows": rows,
        "source_filename": "SOX_Controls_Workbook_2 FINAL.xlsm" if rows else None,
        "imported_at": datetime.now().isoformat(timespec="seconds") if rows else None,
        "imported_by": SEED_IDENTITY if rows else None,
    })
    write_json(pdir / "procedure.json", {
        "columns": [{"key": k, "label": l} for k, l in DEFAULT_PROCEDURE_COLUMNS],
        "steps": procedure_steps,
    })

    echantillons = []
    if echantillons_ids:
        by_id_key = columns[0][0]
        for target_id in echantillons_ids:
            match = next((r for r in rows if str(r.get(by_id_key)) == str(target_id)), None)
            if match:
                echantillons.append({"row": match, "resultat": "A tester", "commentaire": ""})

    control_doc = {
        "date_test": date_test,
        "nb_echantillons": len(echantillons) or 2,
        "echantillons": echantillons,
        "conclusion": "",
        "observations": "",
        "realized_by": SEED_IDENTITY if echantillons else None,
        "validated_by": None,
        "statut_override": None,
    }
    write_json(pdir / "control.json", control_doc)


def main():
    wb = openpyxl.load_workbook(XLSM, data_only=True, keep_vba=True)

    write_json(DATA / "registry.json", {
        "campaign": "Campagne SOX 2026",
        "team": "Tresorerie - Controle Interne",
        "controls": [c["ref"] for c in CONTROLS],
    })
    write_json(DATA / "state.json", {"current_period": PERIOD})

    for c in CONTROLS:
        ref = c["ref"]

        if ref == "N3":
            ws = wb["DATA_N3"]
            rows = [row_to_dict([cell.value for cell in row[1:11]], N3_COLUMNS)
                    for row in ws.iter_rows(min_row=8) if row[1].value]
            seed_control(c, N3_COLUMNS, rows, N3_PROCEDURE,
                         echantillons_ids=["OD-022", "OD-033"], date_test="2026-08-17")

        elif ref == "C5":
            ws = wb["DATA_C5"]
            rows = [row_to_dict([cell.value for cell in row[1:14]], C5_COLUMNS)
                    for row in ws.iter_rows(min_row=8) if row[1].value]
            seed_control(c, C5_COLUMNS, rows, DEFAULT_PROCEDURE_STEPS,
                         echantillons_ids=[5972344, 5963161], date_test="2026-08-17")

        else:
            seed_control(c, [], [], DEFAULT_PROCEDURE_STEPS)

    print("Donnees initialisees dans data/ pour la periode", PERIOD)


if __name__ == "__main__":
    main()
