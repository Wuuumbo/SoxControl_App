#!/usr/bin/env python3
"""Serveur local pour l'outil de suivi des controles SOX.

Bibliotheque standard uniquement, sauf pour l'import de fichiers Excel
(raw_data) qui necessite openpyxl : python -m pip install openpyxl

Lancement : python server.py  (ouvre automatiquement le navigateur)
"""
import ctypes
import json
import mimetypes
import os
import random
import re
import threading
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse, parse_qs

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
CONTROLS_DIR = DATA / "controls"
ELC_DATA = ROOT / "elc" / "data"
PORT = 8765

RESULTATS_CONFORMES = {"OK"}
RESULTATS_NON_CONFORMES = {"KO"}
RESULTATS_NEUTRES = {"N/A"}


# ---------------------------------------------------------------------------
# Identite Windows (non falsifiable depuis le navigateur : calculee ici,
# cote serveur, a partir de la session Windows en cours).
# ---------------------------------------------------------------------------

def _windows_display_name():
    try:
        NAME_DISPLAY = 3
        size = ctypes.c_ulong(0)
        ctypes.windll.secur32.GetUserNameExW(NAME_DISPLAY, None, ctypes.byref(size))
        buf = ctypes.create_unicode_buffer(size.value)
        ok = ctypes.windll.secur32.GetUserNameExW(NAME_DISPLAY, buf, ctypes.byref(size))
        if ok and buf.value:
            return buf.value
    except Exception:
        pass
    return None


def whoami():
    username = os.environ.get("USERNAME", "inconnu")
    return {
        "display_name": _windows_display_name() or username,
        "username": username,
        "computer": os.environ.get("COMPUTERNAME", "inconnu"),
    }


def stamp(existing=None):
    """Identite + horodatage de l'action en cours, calcules cote serveur."""
    data = dict(whoami())
    data["at"] = datetime.now().isoformat(timespec="seconds")
    return data


# ---------------------------------------------------------------------------
# Utilitaires fichiers JSON
# ---------------------------------------------------------------------------

def read_json(path: Path, default):
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp.replace(path)


def registry():
    return read_json(DATA / "registry.json", {"campaign": "", "team": "", "controls": []})


def state():
    return read_json(DATA / "state.json", {"current_period": None})


def set_state(patch: dict):
    s = state()
    s.update(patch)
    write_json(DATA / "state.json", s)
    return s


def control_dir(ref: str) -> Path:
    return CONTROLS_DIR / ref


def meta_path(ref: str) -> Path:
    return control_dir(ref) / "meta.json"


def period_dir(ref: str, period: str) -> Path:
    return control_dir(ref) / "periods" / period


def read_meta(ref: str):
    meta = read_json(meta_path(ref), None)
    if meta is not None:
        meta.setdefault("contacts", [])
        meta.setdefault("mail_templates", [])
    return meta


def list_periods(ref: str):
    pdir = control_dir(ref) / "periods"
    if not pdir.exists():
        return []
    return sorted([p.name for p in pdir.iterdir() if p.is_dir()], reverse=True)


def all_periods():
    periods = set()
    for c in registry().get("controls", []):
        periods.update(list_periods(c))
    return sorted(periods, reverse=True)


def empty_raw_data():
    return {"columns": [], "rows": [], "source_filename": None, "imported_at": None, "imported_by": None}


DEFAULT_PROCEDURE_COLUMNS = [
    {"key": "etape", "label": "Etape"},
    {"key": "action", "label": "Action"},
    {"key": "mode_operatoire", "label": "Mode operatoire"},
    {"key": "statut", "label": "Statut"},
    {"key": "date", "label": "Date"},
    {"key": "contact", "label": "Contact"},
]


def empty_procedure():
    return {"columns": [dict(c) for c in DEFAULT_PROCEDURE_COLUMNS], "steps": []}


def empty_control_doc():
    return {
        "date_test": "",
        "nb_echantillons": 2,
        "echantillons": [],
        "conclusion": "",
        "observations": "",
        "realized_by": None,
        "validated_by": None,
        "statut_override": None,
    }


def ensure_period(ref: str, period: str):
    """Cree le dossier de periode pour un controle s'il n'existe pas, en
    reprenant la structure de la procedure de la periode la plus recente
    disponible (donnees brutes et resultats remis a vide : chaque mois
    reste independant, rien ne s'accumule d'un mois sur l'autre)."""
    pdir = period_dir(ref, period)
    if pdir.exists():
        return
    pdir.mkdir(parents=True, exist_ok=True)

    existing = [p for p in list_periods(ref) if p != period]
    if existing:
        latest = existing[0]
        proc = read_json(period_dir(ref, latest) / "procedure.json", empty_procedure())
        steps = [dict(s, statut="-", date="") for s in proc.get("steps", [])]
        columns = proc.get("columns") or [dict(c) for c in DEFAULT_PROCEDURE_COLUMNS]
        write_json(pdir / "procedure.json", {"columns": columns, "steps": steps})
    else:
        write_json(pdir / "procedure.json", empty_procedure())

    write_json(pdir / "raw_data.json", empty_raw_data())
    write_json(pdir / "control.json", empty_control_doc())


def compute_status(control_doc, nb_rows_data, procedure_steps=None):
    procedure_steps = procedure_steps or []
    echs = control_doc.get("echantillons", [])
    override = control_doc.get("statut_override")

    if override:
        # Validation/signature explicite (bouton "Valider OK" / "Valider KO") :
        # c'est le SEUL moyen d'obtenir OK ou KO. Aucune valeur de la
        # procedure (y compris "Conclusion et signature") ne determine plus
        # OK/KO automatiquement.
        statut = override
    elif not echs:
        # Pas encore d'echantillon teste : des qu'une etape de la procedure a
        # ete touchee (peu importe laquelle et sa valeur, OK ou KO n'a aucun
        # impact ici), le controle passe simplement "En cours".
        statut = "En cours" if any((s.get("statut") or "-") != "-" for s in procedure_steps) else "-"
    elif any(e.get("resultat") in (None, "", "A tester") for e in echs):
        statut = "En cours"
    elif any(e.get("resultat") in RESULTATS_NON_CONFORMES for e in echs):
        statut = "KO"
    else:
        statut = "OK"

    conformes = sum(1 for e in echs if e.get("resultat") in RESULTATS_CONFORMES)
    non_conformes = sum(1 for e in echs if e.get("resultat") in RESULTATS_NON_CONFORMES)
    testables = sum(1 for e in echs if e.get("resultat") in RESULTATS_CONFORMES | RESULTATS_NON_CONFORMES)
    taux = round(100.0 * conformes / testables, 1) if testables else None

    return {
        "statut": statut,
        "nb_echantillons": len(echs),
        "nb_conformes": conformes,
        "nb_non_conformes": non_conformes,
        "taux_conformite": taux,
        "nb_lignes_donnees": nb_rows_data,
    }


def bundle_for(ref: str, period: str):
    meta = read_meta(ref)
    if meta is None:
        return None
    pdir = period_dir(ref, period)
    if not pdir.exists():
        return {"meta": meta, "period": period, "exists": False}
    raw = read_json(pdir / "raw_data.json", empty_raw_data())
    proc = read_json(pdir / "procedure.json", empty_procedure())
    ctrl = read_json(pdir / "control.json", empty_control_doc())
    status = compute_status(ctrl, len(raw.get("rows", [])), proc.get("steps", []))
    return {
        "meta": meta,
        "period": period,
        "exists": True,
        "raw_data": raw,
        "procedure": proc,
        "control": ctrl,
        "status": status,
    }


# ---------------------------------------------------------------------------
# ELC (Entity Level Controls) : un grand tableau segmente par categorie
# (bloc), colonnes et lignes entierement editables.
# ---------------------------------------------------------------------------

ELC_STATUT_OBTENU = {"Obtenu"}
ELC_YEARS_DIR = ELC_DATA / "years"
# Colonnes "de suivi" reinitialisees quand on demarre une nouvelle annee ;
# tout le reste (code ELC, controle, document, equipe, lien racine, statuts
# historiques...) est repris tel quel d'une annee sur l'autre.
ELC_TRACKING_KEYS = {
    "cac_transmission", "date_envoi_mail_cac", "a_presenter_sur_site",
    "commentaire", "statut", "date_obtention", "personne_suivi",
    "commentaires_cac_finale", "lien_fichier",
}


def elc_default_meta():
    return {"categories": [], "columns": [
        {"key": "elc_code", "label": "N. ELC"},
        {"key": "controle", "label": "Controle"},
        {"key": "document", "label": "Document attendu"},
        {"key": "team", "label": "Equipe"},
        {"key": "statut", "label": "Statut"},
        {"key": "commentaire", "label": "Commentaire"},
    ]}


def elc_state():
    return read_json(ELC_DATA / "state.json", {"current_year": None})


def elc_set_state(patch: dict):
    s = elc_state()
    s.update(patch)
    write_json(ELC_DATA / "state.json", s)
    return s


def elc_list_years():
    if not ELC_YEARS_DIR.exists():
        return []
    return sorted([p.name for p in ELC_YEARS_DIR.iterdir() if p.is_dir()], reverse=True)


def elc_year_dir(year):
    return ELC_YEARS_DIR / year


def elc_meta_path(year):
    return elc_year_dir(year) / "meta.json"


def elc_meta(year):
    return read_json(elc_meta_path(year), elc_default_meta())


def elc_rows_path(year, cat_id):
    return elc_year_dir(year) / "rows" / f"{cat_id}.json"


def elc_read_rows(year, cat_id):
    return read_json(elc_rows_path(year, cat_id), {"rows": []})


def elc_ensure_year(year):
    """Cree l'annee si elle n'existe pas, en reprenant la structure (blocs,
    colonnes, lignes) de l'annee la plus recente disponible : seules les
    colonnes de suivi (statut, dates, commentaires, liens...) sont remises a
    vide, le reste (code ELC, controle, document, equipe...) est conserve."""
    if elc_meta_path(year).exists():
        return
    existing = [y for y in elc_list_years() if y != year]
    if not existing:
        write_json(elc_meta_path(year), elc_default_meta())
        return
    latest = existing[0]
    meta = elc_meta(latest)
    write_json(elc_meta_path(year), meta)
    for cat in meta.get("categories", []):
        rows = elc_read_rows(latest, cat["id"]).get("rows", [])
        new_rows = []
        for r in rows:
            nr = dict(r)
            for k in ELC_TRACKING_KEYS:
                if k in nr:
                    nr[k] = "-" if k == "statut" else ""
            new_rows.append(nr)
        write_json(elc_rows_path(year, cat["id"]), {"rows": new_rows})


def elc_all_rows(year):
    meta = elc_meta(year)
    out = []
    for cat in meta.get("categories", []):
        rows = elc_read_rows(year, cat["id"]).get("rows", [])
        for r in rows:
            out.append((cat, r))
    return meta, out


def elc_compute_dashboard(year):
    meta, all_rows = elc_all_rows(year)
    total = len(all_rows)
    by_statut = {}
    by_team = {}
    by_cat = {c["id"]: {"id": c["id"], "nom": c["nom"], "total": 0, "obtenu": 0} for c in meta.get("categories", [])}

    for cat, r in all_rows:
        statut = (r.get("statut") or "-").strip() or "-"
        by_statut[statut] = by_statut.get(statut, 0) + 1
        team = (r.get("team") or "-").strip() or "-"
        team_entry = by_team.setdefault(team, {"total": 0, "obtenu": 0})
        team_entry["total"] += 1
        if statut in ELC_STATUT_OBTENU:
            team_entry["obtenu"] += 1
        by_cat[cat["id"]]["total"] += 1
        if statut in ELC_STATUT_OBTENU:
            by_cat[cat["id"]]["obtenu"] += 1

    obtenu_total = sum(1 for _, r in all_rows if (r.get("statut") or "").strip() in ELC_STATUT_OBTENU)
    taux_global = round(100.0 * obtenu_total / total, 1) if total else 0

    par_categorie = []
    for c in meta.get("categories", []):
        entry = by_cat[c["id"]]
        entry["taux"] = round(100.0 * entry["obtenu"] / entry["total"], 1) if entry["total"] else 0
        par_categorie.append(entry)

    par_equipe = []
    for team, v in sorted(by_team.items()):
        taux = round(100.0 * v["obtenu"] / v["total"], 1) if v["total"] else 0
        par_equipe.append({"team": team, "total": v["total"], "obtenu": v["obtenu"], "taux": taux})

    return {
        "total": total,
        "obtenu_total": obtenu_total,
        "taux_global": taux_global,
        "par_statut": [{"statut": k, "nombre": v} for k, v in sorted(by_statut.items(), key=lambda kv: -kv[1])],
        "par_equipe": par_equipe,
        "par_categorie": par_categorie,
    }


def slugify_key(label, index, used):
    base = re.sub(r"[^a-z0-9]+", "_", (label or "").strip().lower()).strip("_")
    if not base:
        base = f"col_{index + 1}"
    key = base
    n = 2
    while key in used:
        key = f"{base}_{n}"
        n += 1
    used.add(key)
    return key


def parse_excel(path: Path):
    """Lit la 1re feuille d'un classeur Excel : 1re ligne = entetes.
    Necessite openpyxl (pip install openpyxl)."""
    import openpyxl  # import tardif : signale clairement si absent

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []

    used = set()
    columns = []
    for i, label in enumerate(header):
        label = str(label) if label is not None else f"Colonne {i + 1}"
        columns.append({"key": slugify_key(label, i, used), "label": label})

    rows = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        row = {}
        for col, v in zip(columns, raw_row):
            if hasattr(v, "isoformat"):
                v = v.isoformat().split("T")[0] if hasattr(v, "date") else v.isoformat()
            row[col["key"]] = v
        rows.append(row)
    return columns, rows


class Handler(BaseHTTPRequestHandler):
    server_version = "SOXControlesApp/1.0"

    def log_message(self, fmt, *args):
        pass  # silencieux dans la console

    # -- helpers ------------------------------------------------------
    def send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_error_json(self, status, message):
        self.send_json({"error": message}, status=status)

    def read_json_body(self):
        length = int(self.headers.get("Content-Length", 0))
        if length == 0:
            return {}
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8"))

    def read_raw_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length) if length else b""

    def serve_static(self, url_path):
        if url_path == "/":
            url_path = "/login.html"
        rel = url_path.lstrip("/")
        if ".." in rel:
            self.send_error_json(400, "chemin invalide")
            return
        fpath = ROOT / rel
        if not fpath.is_file():
            self.send_error_json(404, "fichier introuvable")
            return
        ctype = {
            ".html": "text/html; charset=utf-8",
            ".css": "text/css; charset=utf-8",
            ".js": "application/javascript; charset=utf-8",
            ".json": "application/json; charset=utf-8",
            ".ico": "image/x-icon",
            ".svg": "image/svg+xml",
        }.get(fpath.suffix) or mimetypes.guess_type(str(fpath))[0] or "application/octet-stream"
        body = fpath.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Content-Disposition", f'attachment; filename="{fpath.name}"' if fpath.suffix not in (".html", ".css", ".js") else "inline")
        self.end_headers()
        self.wfile.write(body)

    # -- routing --------------------------------------------------------
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)
        try:
            if path == "/api/registry":
                reg = registry()
                st = state()
                self.send_json({**reg, "current_period": st.get("current_period")})
                return
            if path == "/api/periods":
                self.send_json({"periods": all_periods(), "current_period": state().get("current_period")})
                return

            if path == "/api/whoami":
                self.send_json(whoami())
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/periods", path)
            if m:
                self.send_json({"periods": list_periods(m.group(1))})
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/([0-9]{4}-[0-9]{2})", path)
            if m:
                ref, period = m.group(1), m.group(2)
                data = bundle_for(ref, period)
                if data is None:
                    self.send_error_json(404, "controle inconnu")
                    return
                self.send_json(data)
                return

            if path == "/api/dashboard":
                period = qs.get("period", [None])[0] or state().get("current_period")
                reg = registry()
                rows = []
                for ref in reg.get("controls", []):
                    meta = read_meta(ref)
                    if meta is None:
                        continue
                    if period and (control_dir(ref) / "periods" / period).exists():
                        pdir = period_dir(ref, period)
                        raw = read_json(pdir / "raw_data.json", empty_raw_data())
                        ctrl = read_json(pdir / "control.json", empty_control_doc())
                        proc = read_json(pdir / "procedure.json", empty_procedure())
                        status = compute_status(ctrl, len(raw.get("rows", [])), proc.get("steps", []))
                    else:
                        status = {"statut": "-", "nb_echantillons": 0, "nb_conformes": 0,
                                  "nb_non_conformes": 0, "taux_conformite": None, "nb_lignes_donnees": 0}
                    rows.append({"ref": ref, "meta": meta, "status": status})
                self.send_json({"period": period, "rows": rows})
                return

            if path == "/api/yearly_status":
                years_available = sorted({p[:4] for p in all_periods()})
                year = qs.get("year", [None])[0] or (years_available[-1] if years_available else str(datetime.now().year))
                refs = registry().get("controls", [])
                months = []
                for m in range(1, 13):
                    period = f"{year}-{m:02d}"
                    counts = {"OK": 0, "KO": 0, "En cours": 0, "-": 0}
                    for ref in refs:
                        pdir = period_dir(ref, period)
                        if pdir.exists():
                            ctrl = read_json(pdir / "control.json", empty_control_doc())
                            raw = read_json(pdir / "raw_data.json", empty_raw_data())
                            proc = read_json(pdir / "procedure.json", empty_procedure())
                            statut = compute_status(ctrl, len(raw.get("rows", [])), proc.get("steps", []))["statut"]
                        else:
                            statut = "-"
                        counts[statut] = counts.get(statut, 0) + 1
                    months.append({"period": period, "counts": counts})
                self.send_json({"year": year, "years_available": years_available, "months": months})
                return

            if path == "/api/elc/years":
                self.send_json({"years": elc_list_years(), "current_year": elc_state().get("current_year")})
                return

            m = re.fullmatch(r"/api/elc/([0-9]{4})/meta", path)
            if m:
                elc_ensure_year(m.group(1))
                self.send_json(elc_meta(m.group(1)))
                return

            m = re.fullmatch(r"/api/elc/([0-9]{4})/dashboard", path)
            if m:
                elc_ensure_year(m.group(1))
                self.send_json(elc_compute_dashboard(m.group(1)))
                return

            m = re.fullmatch(r"/api/elc/([0-9]{4})/categories/([A-Za-z0-9_-]+)/rows", path)
            if m:
                self.send_json(elc_read_rows(m.group(1), m.group(2)))
                return

            # fichiers statiques (html/css/js) + fichiers sources importes (data/...)
            self.serve_static(path)
        except Exception as exc:  # noqa: BLE001
            self.send_error_json(500, str(exc))

    def do_PUT(self):
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            body = self.read_json_body()

            if path == "/api/state":
                self.send_json(set_state(body))
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/meta", path)
            if m:
                ref = m.group(1)
                if meta_path(ref).exists():
                    body.setdefault("contacts", [])
                    body.setdefault("mail_templates", [])
                    write_json(meta_path(ref), body)
                    self.send_json(body)
                else:
                    self.send_error_json(404, "controle inconnu")
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/([0-9]{4}-[0-9]{2})/procedure", path)
            if m:
                ref, period = m.group(1), m.group(2)
                ensure_period(ref, period)
                body.setdefault("columns", [dict(c) for c in DEFAULT_PROCEDURE_COLUMNS])
                write_json(period_dir(ref, period) / "procedure.json", body)
                ctrl = read_json(period_dir(ref, period) / "control.json", empty_control_doc())
                raw = read_json(period_dir(ref, period) / "raw_data.json", empty_raw_data())
                status = compute_status(ctrl, len(raw.get("rows", [])), body.get("steps", []))
                self.send_json({"ok": True, "saved_at": datetime.now().isoformat(timespec="seconds"), "status": status})
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/([0-9]{4}-[0-9]{2})/control", path)
            if m:
                ref, period = m.group(1), m.group(2)
                ensure_period(ref, period)
                # L'identite "realise par" et le statut de validation sont
                # calcules/preserves cote serveur : le client ne peut pas les
                # falsifier via le corps de la requete.
                previous = read_json(period_dir(ref, period) / "control.json", empty_control_doc())
                body["realized_by"] = stamp()
                body["validated_by"] = previous.get("validated_by")
                body["statut_override"] = previous.get("statut_override")
                write_json(period_dir(ref, period) / "control.json", body)
                raw = read_json(period_dir(ref, period) / "raw_data.json", empty_raw_data())
                proc = read_json(period_dir(ref, period) / "procedure.json", empty_procedure())
                status = compute_status(body, len(raw.get("rows", [])), proc.get("steps", []))
                self.send_json({"ok": True, "saved_at": datetime.now().isoformat(timespec="seconds"),
                                 "status": status, "control": body})
                return

            if path == "/api/elc/state":
                self.send_json(elc_set_state(body))
                return

            m = re.fullmatch(r"/api/elc/([0-9]{4})/meta", path)
            if m:
                write_json(elc_meta_path(m.group(1)), body)
                self.send_json(body)
                return

            m = re.fullmatch(r"/api/elc/([0-9]{4})/categories/([A-Za-z0-9_-]+)/rows", path)
            if m:
                write_json(elc_rows_path(m.group(1), m.group(2)), body)
                self.send_json({"ok": True, "saved_at": datetime.now().isoformat(timespec="seconds")})
                return

            self.send_error_json(404, "route inconnue")
        except Exception as exc:  # noqa: BLE001
            self.send_error_json(500, str(exc))

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/([0-9]{4}-[0-9]{2})/import_raw_data", path)
            if m:
                ref, period = m.group(1), m.group(2)
                ensure_period(ref, period)
                filename = self.headers.get("X-Filename", "import.xlsx")
                safe_name = re.sub(r"[^A-Za-z0-9._ -]", "_", filename).strip() or "import.xlsx"
                raw_bytes = self.read_raw_body()
                if not raw_bytes:
                    self.send_error_json(400, "fichier vide")
                    return

                src_dir = period_dir(ref, period) / "source"
                if src_dir.exists():
                    for old in src_dir.iterdir():
                        old.unlink()
                src_dir.mkdir(parents=True, exist_ok=True)
                dest = src_dir / safe_name
                dest.write_bytes(raw_bytes)

                try:
                    columns, rows = parse_excel(dest)
                except ModuleNotFoundError:
                    self.send_error_json(500, "openpyxl n'est pas installe : lancez 'python -m pip install openpyxl'")
                    return
                except Exception as exc:  # noqa: BLE001
                    self.send_error_json(400, f"impossible de lire ce fichier Excel : {exc}")
                    return

                raw_data = {
                    "columns": columns,
                    "rows": rows,
                    "source_filename": safe_name,
                    "imported_at": datetime.now().isoformat(timespec="seconds"),
                    "imported_by": whoami(),
                }
                write_json(period_dir(ref, period) / "raw_data.json", raw_data)

                ctrl = read_json(period_dir(ref, period) / "control.json", empty_control_doc())
                proc = read_json(period_dir(ref, period) / "procedure.json", empty_procedure())
                status = compute_status(ctrl, len(rows), proc.get("steps", []))
                self.send_json({"ok": True, "raw_data": raw_data, "status": status})
                return

            body = self.read_json_body()

            if path == "/api/periods":
                period = body.get("period")
                if not period or not re.fullmatch(r"[0-9]{4}-[0-9]{2}", period):
                    self.send_error_json(400, "periode invalide (attendu AAAA-MM)")
                    return
                for ref in registry().get("controls", []):
                    ensure_period(ref, period)
                set_state({"current_period": period})
                self.send_json({"ok": True, "period": period})
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/([0-9]{4}-[0-9]{2})/draw_samples", path)
            if m:
                ref, period = m.group(1), m.group(2)
                ensure_period(ref, period)
                n = int(body.get("n") or 2)
                raw = read_json(period_dir(ref, period) / "raw_data.json", empty_raw_data())
                rows = raw.get("rows", [])
                if not rows:
                    self.send_error_json(400, "aucune donnee brute importee pour tirer un echantillon")
                    return
                n = min(n, len(rows))
                picked = random.sample(rows, n)
                ctrl = read_json(period_dir(ref, period) / "control.json", empty_control_doc())
                ctrl["nb_echantillons"] = n
                ctrl["echantillons"] = [
                    {"row": row, "resultat": "A tester", "commentaire": ""} for row in picked
                ]
                ctrl["realized_by"] = stamp()
                write_json(period_dir(ref, period) / "control.json", ctrl)
                proc = read_json(period_dir(ref, period) / "procedure.json", empty_procedure())
                status = compute_status(ctrl, len(rows), proc.get("steps", []))
                self.send_json({"ok": True, "control": ctrl, "status": status})
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/([0-9]{4}-[0-9]{2})/validate", path)
            if m:
                ref, period = m.group(1), m.group(2)
                ensure_period(ref, period)
                statut_signe = body.get("statut") if body.get("statut") in ("OK", "KO") else "OK"
                ctrl = read_json(period_dir(ref, period) / "control.json", empty_control_doc())
                ctrl["validated_by"] = stamp()
                # Valider/signer un controle le cloture systematiquement au
                # statut choisi (OK ou KO) : c'est l'engagement du controleur.
                ctrl["statut_override"] = statut_signe
                write_json(period_dir(ref, period) / "control.json", ctrl)
                raw = read_json(period_dir(ref, period) / "raw_data.json", empty_raw_data())
                proc = read_json(period_dir(ref, period) / "procedure.json", empty_procedure())
                status = compute_status(ctrl, len(raw.get("rows", [])), proc.get("steps", []))
                self.send_json({"ok": True, "control": ctrl, "status": status})
                return

            m = re.fullmatch(r"/api/controls/([A-Za-z0-9]+)/([0-9]{4}-[0-9]{2})/invalidate", path)
            if m:
                ref, period = m.group(1), m.group(2)
                ensure_period(ref, period)
                ctrl = read_json(period_dir(ref, period) / "control.json", empty_control_doc())
                ctrl["validated_by"] = None
                ctrl["statut_override"] = None
                write_json(period_dir(ref, period) / "control.json", ctrl)
                raw = read_json(period_dir(ref, period) / "raw_data.json", empty_raw_data())
                proc = read_json(period_dir(ref, period) / "procedure.json", empty_procedure())
                status = compute_status(ctrl, len(raw.get("rows", [])), proc.get("steps", []))
                self.send_json({"ok": True, "control": ctrl, "status": status})
                return

            if path == "/api/elc/years":
                year = (body.get("year") or "").strip()
                if not re.fullmatch(r"[0-9]{4}", year):
                    self.send_error_json(400, "annee invalide (attendu AAAA)")
                    return
                elc_ensure_year(year)
                self.send_json(elc_set_state({"current_year": year}))
                return

            m = re.fullmatch(r"/api/elc/([0-9]{4})/categories", path)
            if m:
                year = m.group(1)
                nom = (body.get("nom") or "").strip()
                if not nom:
                    self.send_error_json(400, "nom de bloc requis")
                    return
                meta = elc_meta(year)
                new_id = re.sub(r"[^a-z0-9]+", "-", nom.lower()).strip("-") or f"bloc-{len(meta['categories']) + 1}"
                base_id = new_id
                n = 2
                existing_ids = {c["id"] for c in meta["categories"]}
                while new_id in existing_ids:
                    new_id = f"{base_id}-{n}"
                    n += 1
                meta["categories"].append({"id": new_id, "nom": nom})
                write_json(elc_meta_path(year), meta)
                write_json(elc_rows_path(year, new_id), {"rows": []})
                self.send_json(meta)
                return

            m = re.fullmatch(r"/api/elc/([0-9]{4})/categories/([A-Za-z0-9_-]+)/delete", path)
            if m:
                year, cat_id = m.group(1), m.group(2)
                meta = elc_meta(year)
                meta["categories"] = [c for c in meta["categories"] if c["id"] != cat_id]
                write_json(elc_meta_path(year), meta)
                rp = elc_rows_path(year, cat_id)
                if rp.exists():
                    rp.unlink()
                self.send_json(meta)
                return

            self.send_error_json(404, "route inconnue")
        except Exception as exc:  # noqa: BLE001
            self.send_error_json(500, str(exc))


def main():
    server = ThreadingHTTPServer(("127.0.0.1", PORT), Handler)
    url = f"http://127.0.0.1:{PORT}/"
    print(f"Suivi des controles SOX - serveur demarre sur {url}")
    print(f"Identifie comme : {whoami()['display_name']} ({whoami()['computer']})")
    print("Ctrl+C pour arreter.")
    threading.Timer(0.6, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nArret du serveur.")


if __name__ == "__main__":
    main()
